"""
Flask Backend for BLE Tag Tracker
Handles OAuth token fetching to bypass CORS restrictions
Production-ready with environment variable configuration
"""

from flask import Flask, jsonify, send_from_directory, make_response, request
from flask_cors import CORS
import requests
import os
import jwt
import random
import smtplib
import hashlib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from functools import wraps
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

app = Flask(__name__, static_folder='../.')

# Configure CORS for production
CORS(app, resources={r"/api/*": {"origins": "*"}})

# Secret key for JWT
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'your-secret-key-change-in-production-' + str(random.randint(1000, 9999)))

# Email Configuration (for PIN sending)
EMAIL_HOST = os.getenv('EMAIL_HOST', 'smtp.gmail.com')
EMAIL_PORT = int(os.getenv('EMAIL_PORT', 587))
EMAIL_USERNAME = os.getenv('EMAIL_USERNAME', '')
EMAIL_PASSWORD = os.getenv('EMAIL_PASSWORD', '')
EMAIL_FROM = os.getenv('EMAIL_FROM', EMAIL_USERNAME)

# OAuth Configuration from environment variables
TOKEN_URL = os.getenv('TOKEN_URL', 'https://login.mzoneweb.net/connect/token')
CLIENT_ID = os.getenv('CLIENT_ID')
CLIENT_SECRET = os.getenv('CLIENT_SECRET')
OAUTH_USERNAME = os.getenv('OAUTH_USERNAME')
OAUTH_PASSWORD = os.getenv('OAUTH_PASSWORD')
SCOPE = os.getenv('SCOPE', 'mz6-api.all mz_username')
GRANT_TYPE = os.getenv('GRANT_TYPE', 'password')
RESPONSE_TYPE = os.getenv('RESPONSE_TYPE', 'code id_token')

# Server Configuration
DEBUG_MODE = os.getenv('DEBUG', 'False').lower() == 'true'
HOST = os.getenv('HOST', '0.0.0.0')
PORT = int(os.getenv('PORT', 5000))

# Check if OAuth credentials are configured (not required for admin portal)
OAUTH_CONFIGURED = all([CLIENT_ID, CLIENT_SECRET, OAUTH_USERNAME, OAUTH_PASSWORD])

# Token cache
token_cache = {
    'token': None,
    'expires_at': None
}

# PIN storage (in production, use Redis or database)
pin_storage = {}

# User storage (in production, use a proper database)
users_db = {}

# Admin credentials (from environment variables)
ADMIN_USERNAME = os.getenv('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.getenv('ADMIN_PASSWORD', 'changeme')

# Billing configuration
billing_config = {
    'cost_per_user': 10.0,  # Default $10 per user per month
    'currency': 'USD'
}

# IMEI tracking (linked to users)
user_imeis = {}  # Format: {email: [{imei: 'xxx', added_date: 'iso-date'}, ...]}


# ============================================================================
# AUTHENTICATION HELPERS
# ============================================================================

def generate_pin():
    """Generate a random 6-digit PIN"""
    return "123456"  # DEMO: Hardcoded for tonight's demo
    # return str(random.randint(100000, 999999))  # Original random PIN


def send_email_pin(email, pin):
    """Send PIN via email"""
    try:
        # If no email config, just print PIN (development mode)
        if not EMAIL_USERNAME or not EMAIL_PASSWORD:
            print(f"\n{'='*60}", flush=True)
            print(f"📧 EMAIL PIN for {email}: {pin}", flush=True)
            print(f"{'='*60}\n", flush=True)
            return True
        
        # Create message
        msg = MIMEMultipart('alternative')
        msg['Subject'] = 'Your BLE Tracker Login Code'
        msg['From'] = EMAIL_FROM
        msg['To'] = email
        
        # HTML email body
        html = f"""
        <html>
          <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px;">
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; border-radius: 10px; text-align: center;">
              <h1 style="color: white; margin: 0;">BLE Tag Tracker</h1>
            </div>
            <div style="background: #f7f7f7; padding: 30px; border-radius: 10px; margin-top: 20px;">
              <h2 style="color: #333;">Your Verification Code</h2>
              <p style="color: #666; font-size: 16px;">Use this code to complete your login:</p>
              <div style="background: white; padding: 20px; border-radius: 8px; text-align: center; margin: 20px 0;">
                <span style="font-size: 36px; font-weight: bold; color: #667eea; letter-spacing: 8px;">{pin}</span>
              </div>
              <p style="color: #666; font-size: 14px;">This code will expire in 10 minutes.</p>
              <p style="color: #999; font-size: 12px; margin-top: 30px;">If you didn't request this code, please ignore this email.</p>
            </div>
          </body>
        </html>
        """
        
        text = f"Your BLE Tracker verification code is: {pin}\n\nThis code will expire in 10 minutes."
        
        part1 = MIMEText(text, 'plain')
        part2 = MIMEText(html, 'html')
        msg.attach(part1)
        msg.attach(part2)
        
        # Send email
        with smtplib.SMTP(EMAIL_HOST, EMAIL_PORT) as server:
            server.starttls()
            server.login(EMAIL_USERNAME, EMAIL_PASSWORD)
            server.send_message(msg)
        
        if DEBUG_MODE:
            print(f"✅ PIN email sent to {email}")
        
        return True
        
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Failed to send email: {str(e)}")
            print(f"📧 Development Mode - PIN for {email}: {pin}")
        return True  # Return True in development mode


def generate_jwt_token(email):
    """Generate JWT token for authenticated user"""
    payload = {
        'email': email,
        'exp': datetime.utcnow() + timedelta(days=30),  # Token valid for 30 days
        'iat': datetime.utcnow()
    }
    return jwt.encode(payload, app.config['SECRET_KEY'], algorithm='HS256')


def verify_jwt_token(token):
    """Verify JWT token and return email"""
    try:
        payload = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
        return payload['email']
    except jwt.ExpiredSignatureError:
        return None
    except jwt.InvalidTokenError:
        return None


def token_required(f):
    """Decorator to require JWT token for protected routes"""
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        
        # Check for token in Authorization header
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(' ')[1]  # Bearer <token>
            except IndexError:
                return jsonify({'error': 'Invalid token format'}), 401
        
        if not token:
            return jsonify({'error': 'Token is missing'}), 401
        
        # Verify token
        email = verify_jwt_token(token)
        if not email:
            return jsonify({'error': 'Token is invalid or expired'}), 401
        
        return f(email, *args, **kwargs)
    
    return decorated


# ============================================================================
# AUTHENTICATION ROUTES
# ============================================================================

@app.route('/api/auth/send-pin', methods=['POST'])
def send_pin():
    """Send PIN to user's email"""
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        is_new_user = data.get('isNewUser', False)
        
        if not email:
            return jsonify({'error': 'Email is required'}), 400
        
        # Basic email validation
        if '@' not in email or '.' not in email:
            return jsonify({'error': 'Invalid email address'}), 400
        
        # Check if user exists for sign-in
        if not is_new_user and email not in users_db:
            return jsonify({'error': 'User not found. Please use Get Started instead.'}), 404
        
        # Generate PIN
        pin = generate_pin()
        
        # Store PIN with expiration (10 minutes)
        pin_storage[email] = {
            'pin': pin,
            'expires_at': datetime.now() + timedelta(minutes=10),
            'attempts': 0
        }
        
        # Send PIN via email
        if send_email_pin(email, pin):
            if DEBUG_MODE:
                print(f"🔐 PIN generated for {email}: {pin}", flush=True)
            
            return jsonify({
                'success': True,
                'message': 'Verification code sent to your email'
            })
        else:
            return jsonify({'error': 'Failed to send email'}), 500
            
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Error in send_pin: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/auth/verify-pin', methods=['POST'])
def verify_pin():
    """Verify PIN and return JWT token"""
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        pin = data.get('pin', '').strip()
        
        if not email or not pin:
            return jsonify({'error': 'Email and PIN are required'}), 400
        
        # Check if PIN exists and is not expired
        if email not in pin_storage:
            return jsonify({'error': 'No verification code found. Please request a new one.'}), 404
        
        pin_data = pin_storage[email]
        
        # Check expiration
        if datetime.now() > pin_data['expires_at']:
            del pin_storage[email]
            return jsonify({'error': 'Verification code expired. Please request a new one.'}), 400
        
        # Check attempts (max 5)
        if pin_data['attempts'] >= 5:
            del pin_storage[email]
            return jsonify({'error': 'Too many failed attempts. Please request a new code.'}), 429
        
        # Verify PIN
        if pin != pin_data['pin']:
            pin_data['attempts'] += 1
            return jsonify({'error': 'Invalid verification code'}), 401
        
        # PIN is valid - create/update user
        if email not in users_db:
            users_db[email] = {
                'email': email,
                'created_at': datetime.now().isoformat(),
                'last_login': datetime.now().isoformat()
            }
        else:
            users_db[email]['last_login'] = datetime.now().isoformat()
        
        # Clean up PIN
        del pin_storage[email]
        
        # Generate JWT token
        token = generate_jwt_token(email)
        
        if DEBUG_MODE:
            print(f"✅ User {email} authenticated successfully")
        
        return jsonify({
            'success': True,
            'token': token,
            'user': {
                'email': email
            }
        })
        
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Error in verify_pin: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/auth/verify-token', methods=['GET'])
@token_required
def verify_token_route(email):
    """Verify if token is still valid"""
    return jsonify({
        'success': True,
        'email': email
    })


@app.route('/api/auth/logout', methods=['POST'])
@token_required
def logout(email):
    """Logout user (client should delete token)"""
    if DEBUG_MODE:
        print(f"👋 User {email} logged out")
    
    return jsonify({
        'success': True,
        'message': 'Logged out successfully'
    })


@app.route('/api/auth/update-profile', methods=['POST'])
@token_required
def update_profile(email):
    """Update user profile with name and password"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        first_name = data.get('first_name', '').strip()
        last_name = data.get('last_name', '').strip()
        password = data.get('password', '').strip()
        
        # Validate inputs
        if not first_name or not last_name:
            return jsonify({'error': 'First name and last name are required'}), 400
        
        if not password or len(password) < 8:
            return jsonify({'error': 'Password must be at least 8 characters'}), 400
        
        # Hash password using SHA-256
        password_hash = hashlib.sha256(password.encode()).hexdigest()
        
        # Update user in database
        if email not in users_db:
            return jsonify({'error': 'User not found'}), 404
        
        users_db[email]['first_name'] = first_name
        users_db[email]['last_name'] = last_name
        users_db[email]['password_hash'] = password_hash
        users_db[email]['profile_completed'] = True
        users_db[email]['updated_at'] = datetime.now().isoformat()
        
        if DEBUG_MODE:
            print(f"✅ Profile updated for {email}: {first_name} {last_name}")
        
        return jsonify({
            'success': True,
            'message': 'Profile updated successfully',
            'user': {
                'email': email,
                'first_name': first_name,
                'last_name': last_name
            }
        })
        
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Error in update_profile: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500


# ============================================================================
# STATIC FILE ROUTES
# ============================================================================

@app.route('/')
def index():
    """Serve the login page by default"""
    return send_from_directory('../', 'login.html')


@app.route('/manifest.json')
def manifest():
    """Serve PWA manifest with correct MIME type"""
    response = make_response(send_from_directory('../', 'manifest.json'))
    response.headers['Content-Type'] = 'application/manifest+json'
    return response


@app.route('/service-worker.js')
def service_worker():
    """Serve service worker with correct MIME type and no caching"""
    response = make_response(send_from_directory('../', 'service-worker.js'))
    response.headers['Content-Type'] = 'application/javascript'
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/<path:path>')
def serve_static(path):
    """Serve static files (CSS, JS, images)"""
    return send_from_directory('../', path)


@app.route('/api/validate-imei/<imei>', methods=['GET'])
@token_required
def validate_imei(current_user, imei):
    """
    Validate IMEI by checking if it exists in the MProfiler API
    Returns 200 if valid, 404 if not found, 400 if error
    """
    try:
        # Validate IMEI format (should be 15 digits)
        if not imei.isdigit() or len(imei) != 15:
            return jsonify({
                'success': False,
                'error': 'Invalid IMEI format. Must be 15 digits.'
            }), 400
        
        # Call MProfiler API to validate IMEI
        validation_url = f'https://live.scopemp.net/Scope.MProfiler.Api/v1/UnitExtendedProperties/{imei}'
        headers = {
            'Authorization': 'Basic T0NJX1Njb3BlVUtfTVo6T0NJU2NvcGVVS01aMDEh'
        }
        
        if DEBUG_MODE:
            print(f"🔍 Validating IMEI: {imei}")
            print(f"📡 API URL: {validation_url}")
        
        response = requests.get(validation_url, headers=headers, timeout=10)
        
        if response.status_code == 200:
            if DEBUG_MODE:
                print(f"✅ IMEI {imei} is valid")
            return jsonify({
                'success': True,
                'message': 'IMEI is valid',
                'imei': imei
            }), 200
        elif response.status_code == 404:
            if DEBUG_MODE:
                print(f"❌ IMEI {imei} not found")
            return jsonify({
                'success': False,
                'error': 'IMEI unknown error'
            }), 404
        else:
            if DEBUG_MODE:
                print(f"⚠️ IMEI validation failed with status: {response.status_code}")
            return jsonify({
                'success': False,
                'error': f'Validation service error: {response.status_code}'
            }), 500
            
    except requests.exceptions.Timeout:
        if DEBUG_MODE:
            print(f"⏱️ IMEI validation timeout for {imei}")
        return jsonify({
            'success': False,
            'error': 'Validation service timeout. Please try again.'
        }), 500
    except requests.exceptions.RequestException as e:
        if DEBUG_MODE:
            print(f"❌ IMEI validation error: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Unable to validate IMEI. Please try again.'
        }), 500
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Unexpected error during IMEI validation: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'An unexpected error occurred'
        }), 500


@app.route('/api/token', methods=['GET'])
def get_token():
    """
    Fetch OAuth token from login.mzoneweb.net
    Returns cached token if still valid, otherwise fetches new one
    """
    try:
        # Check if OAuth is configured
        if not OAUTH_CONFIGURED:
            return jsonify({
                'success': False,
                'error': 'OAuth not configured. Please set credentials in .env file'
            }), 503
        
        # Check if cached token is still valid (with 5 min buffer)
        if token_cache['token'] and token_cache['expires_at']:
            if datetime.now() < token_cache['expires_at'] - timedelta(minutes=5):
                return jsonify({
                    'success': True,
                    'access_token': token_cache['token'],
                    'expires_in': int((token_cache['expires_at'] - datetime.now()).total_seconds()),
                    'cached': True
                })
        
        # Fetch new token
        if DEBUG_MODE:
            print("🔐 Fetching new OAuth token...")
        
        # Build OAuth payload from environment variables
        # Using stored credentials for API access
        from urllib.parse import quote
        
        payload = {
            'client_id': CLIENT_ID,
            'client_secret': CLIENT_SECRET,
            'username': OAUTH_USERNAME,
            'password': OAUTH_PASSWORD,
            'scope': 'mz6-api.all mz_username',
            'grant_type': GRANT_TYPE,
            'response_type': 'code id token'
        }
        
        # URL encode the payload
        payload_str = '&'.join([f"{k}={quote(str(v))}" for k, v in payload.items()])
        
        if DEBUG_MODE:
            print(f"📤 Sending payload to: {TOKEN_URL}")
            print(f"📋 Using OAuth credentials from environment")
        
        response = requests.post(
            TOKEN_URL,
            data=payload_str,
            headers={'Content-Type': 'application/x-www-form-urlencoded'},
            timeout=10
        )
        
        if response.status_code == 200:
            data = response.json()
            
            # Cache the token
            token_cache['token'] = data['access_token']
            expires_in = data.get('expires_in', 3600)
            token_cache['expires_at'] = datetime.now() + timedelta(seconds=expires_in)
            
            if DEBUG_MODE:
                print(f"✅ Token obtained (expires in {expires_in}s)")
            
            return jsonify({
                'success': True,
                'access_token': data['access_token'],
                'expires_in': expires_in,
                'cached': False
            })
        else:
            error_detail = response.text
            if DEBUG_MODE:
                print(f"❌ Token fetch failed: {response.status_code}")
                print(f"📋 OAuth Error Response: {error_detail}")
                print(f"📤 Payload sent: client_id={CLIENT_ID}, username={OAUTH_USERNAME}")
            return jsonify({
                'success': False,
                'error': f'Token request failed: {response.status_code}',
                'detail': error_detail
            }), response.status_code
            
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Exception: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Internal server error'
        }), 500


@app.route('/api/vehicles', methods=['POST'])
@token_required
def get_vehicles(current_user):
    """
    Fetch vehicles from mzone API and filter by user's registered IMEIs
    Matches IMEI against 'registration' field in Vehicle data
    Returns: Filtered vehicle list with IDs for fetching LastKnownPosition
    """
    try:
        data = request.get_json()
        imeis = data.get('imeis', [])
        
        if not imeis or not isinstance(imeis, list):
            return jsonify({
                'success': False,
                'error': 'No IMEIs provided. Please add BLE tags first.'
            }), 400
        
        if DEBUG_MODE:
            print(f"🚗 Fetching vehicles for user {current_user}")
            print(f"📋 Registered IMEIs: {imeis}")
        
        # Get valid token
        token_response = get_token()
        if isinstance(token_response, tuple):
            return token_response
        
        token_data = token_response.get_json()
        if not token_data.get('success'):
            return jsonify({
                'success': False,
                'error': 'Failed to obtain access token'
            }), 401
        
        access_token = token_data['access_token']
        
        # Fetch ALL vehicles from mzone API with expanded position data
        # This API includes registration field which holds the IMEI
        api_url = 'https://live.mzoneweb.net/mzone62.api/Vehicles?$format=json&$count=true&$select=id%2Cdescription%2CvehicleIcon%2CvehicleIconColor%2ClastKnownEventUtcLastModified%2CdisablementFeatureAvailable%2CignitionOn%2Cregistration%2ClastKnownEventUtcTimestamp%2ClastKnownGpsEventUtcTimestamp%2CdecimalOdometer%2CengineSeconds%2Cvin%2Cunit_Description%2CvehicleDisabled&$orderby=description&$skip=0&$top=1000&vehicleGroup_Id=8e4fe3a7-4d46-474d-bf57-993828f70968&$expand=lastKnownPosition(%24select%3Dlongitude%2Clatitude%2CeventType_Id%2CutcTimestamp%2Cspeed)%2ClastKnownTemperature(%24select%3DutcTimestamp)%2CvehicleType(%24select%3DvehicleNotReportedThreshold)%2Cmake(%24select%3Ddescription)%2Cmodel(%24select%3Ddescription)%2ClastKnownFuelConsumption(%24select%3DaverageConsumption%2CaverageConsumptionUtcTimestamp%2CinstantaneousConsumption%2CinstantaneousConsumptionUtcTimestamp)%2ClastKnownFuelLevel(%24select%3Dlevel%2Cpercentage%2CutcTimestamp)'
        
        headers = {
            'Authorization': f'Bearer {access_token}'
        }
        
        if DEBUG_MODE:
            print(f"🌐 Fetching all vehicles from API...")
        
        response = requests.get(api_url, headers=headers, timeout=30)
        
        if response.status_code == 200:
            vehicles_data = response.json()
            all_vehicles = vehicles_data.get('value', [])
            
            if DEBUG_MODE:
                print(f"✅ Fetched {len(all_vehicles)} total vehicles")
            
            # Filter vehicles by matching registration field with user's IMEIs
            filtered_vehicles = []
            for vehicle in all_vehicles:
                registration = vehicle.get('registration', '')
                # Match registration against user's IMEIs
                if registration in imeis:
                    filtered_vehicles.append(vehicle)
                    if DEBUG_MODE:
                        vehicle_id = vehicle.get('id', 'N/A')
                        description = vehicle.get('description', 'Unknown')
                        print(f"✅ Matched: {description} (Registration: {registration}, ID: {vehicle_id})")
            
            if DEBUG_MODE:
                print(f"🎯 Filtered to {len(filtered_vehicles)} vehicles matching user IMEIs")
            
            return jsonify({
                'success': True,
                'vehicles': {
                    'value': filtered_vehicles,
                    '@odata.count': len(filtered_vehicles)
                },
                'requested_imeis': len(imeis),
                'found_vehicles': len(filtered_vehicles)
            })
        else:
            if DEBUG_MODE:
                print(f"❌ Failed to fetch vehicles: {response.status_code}")
                print(f"📄 Response: {response.text[:500]}")
            return jsonify({
                'success': False,
                'error': f'API request failed: {response.status_code}'
            }), response.status_code
            
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Exception fetching vehicles: {str(e)}")
            import traceback
            traceback.print_exc()
        return jsonify({
            'success': False,
            'error': 'Internal server error'
        }), 500


@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    return jsonify({
        'status': 'healthy',
        'token_cached': token_cache['token'] is not None,
        'token_expires_at': token_cache['expires_at'].isoformat() if token_cache['expires_at'] else None
    })


# ============================================================================
# ADMIN AUTHENTICATION & BILLING
# ============================================================================

def admin_auth_required(f):
    """Decorator to require admin authentication"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        auth_header = request.headers.get('Authorization', '')
        
        if not auth_header.startswith('Admin '):
            return jsonify({'error': 'Unauthorized'}), 401
        
        try:
            # Extract token (format: "Admin base64(username:password)")
            import base64
            token = auth_header.split(' ')[1]
            credentials = base64.b64decode(token).decode('utf-8')
            username, password = credentials.split(':')
            
            if username != ADMIN_USERNAME or password != ADMIN_PASSWORD:
                return jsonify({'error': 'Invalid credentials'}), 401
            
            return f(*args, **kwargs)
        except:
            return jsonify({'error': 'Invalid authorization format'}), 401
    
    return decorated_function


@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    """Admin login endpoint"""
    try:
        data = request.get_json()
        username = data.get('username', '')
        password = data.get('password', '')
        
        if username == ADMIN_USERNAME and password == ADMIN_PASSWORD:
            import base64
            # Create a simple token (base64 encoded credentials)
            token = base64.b64encode(f"{username}:{password}".encode()).decode()
            
            return jsonify({
                'success': True,
                'token': token,
                'message': 'Login successful'
            })
        else:
            return jsonify({'error': 'Invalid credentials'}), 401
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Error in admin_login: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/admin/users', methods=['GET'])
@admin_auth_required
def admin_get_users():
    """Get all users with their IMEIs"""
    try:
        users_list = []
        
        for email, user_data in users_db.items():
            imeis = user_imeis.get(email, [])
            
            users_list.append({
                'email': email,
                'created_at': user_data.get('created_at'),
                'last_login': user_data.get('last_login'),
                'imeis': imeis,
                'imei_count': len(imeis)
            })
        
        # Sort by created date (newest first)
        users_list.sort(key=lambda x: x['created_at'], reverse=True)
        
        return jsonify({
            'success': True,
            'users': users_list,
            'total_users': len(users_list)
        })
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Error in admin_get_users: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/admin/user/<email>/imeis', methods=['POST'])
@admin_auth_required
def admin_add_imei(email):
    """Add IMEI to user account"""
    try:
        data = request.get_json()
        imei = data.get('imei', '').strip()
        
        if not imei:
            return jsonify({'error': 'IMEI is required'}), 400
        
        if email not in users_db:
            return jsonify({'error': 'User not found'}), 404
        
        # Initialize IMEI list if needed
        if email not in user_imeis:
            user_imeis[email] = []
        
        # Check if IMEI already exists for this user
        existing = [i for i in user_imeis[email] if i['imei'] == imei]
        if existing:
            return jsonify({'error': 'IMEI already exists for this user'}), 400
        
        # Add IMEI
        user_imeis[email].append({
            'imei': imei,
            'added_date': datetime.now().isoformat()
        })
        
        return jsonify({
            'success': True,
            'message': 'IMEI added successfully'
        })
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Error in admin_add_imei: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/admin/user/<email>/imeis/<imei>', methods=['DELETE'])
@admin_auth_required
def admin_remove_imei(email, imei):
    """Remove IMEI from user account"""
    try:
        if email not in user_imeis:
            return jsonify({'error': 'User has no IMEIs'}), 404
        
        # Find and remove IMEI
        user_imeis[email] = [i for i in user_imeis[email] if i['imei'] != imei]
        
        return jsonify({
            'success': True,
            'message': 'IMEI removed successfully'
        })
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Error in admin_remove_imei: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/admin/billing/config', methods=['GET'])
@admin_auth_required
def admin_get_billing_config():
    """Get billing configuration"""
    return jsonify({
        'success': True,
        'config': billing_config
    })


@app.route('/api/admin/billing/config', methods=['PUT'])
@admin_auth_required
def admin_update_billing_config():
    """Update billing configuration"""
    try:
        data = request.get_json()
        cost = data.get('cost_per_user')
        
        if cost is not None:
            billing_config['cost_per_user'] = float(cost)
        
        if 'currency' in data:
            billing_config['currency'] = data['currency']
        
        return jsonify({
            'success': True,
            'config': billing_config,
            'message': 'Billing configuration updated'
        })
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Error in admin_update_billing_config: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/admin/billing/calculate', methods=['GET'])
@admin_auth_required
def admin_calculate_billing():
    """Calculate monthly billing"""
    try:
        total_users = len(users_db)
        cost_per_user = billing_config['cost_per_user']
        total_cost = total_users * cost_per_user
        
        return jsonify({
            'success': True,
            'billing': {
                'total_users': total_users,
                'cost_per_user': cost_per_user,
                'currency': billing_config['currency'],
                'total_monthly_cost': total_cost,
                'calculation_date': datetime.now().isoformat()
            }
        })
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Error in admin_calculate_billing: {str(e)}")
        return jsonify({'error': 'Internal server error'}), 500


@app.route('/api/admin/billing/export', methods=['GET'])
@admin_auth_required
def admin_export_billing():
    """Export billing statement to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Billing Statement"
        
        # Header styling
        header_fill = PatternFill(start_color="173D64", end_color="173D64", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = 'BLE Tag Tracker - Monthly Billing Statement'
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Date
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Empty row
        ws.append([])
        
        # Summary
        total_users = len(users_db)
        cost_per_user = billing_config['cost_per_user']
        total_cost = total_users * cost_per_user
        
        ws.append(['Summary'])
        ws['A4'].font = Font(bold=True, size=14)
        ws.append(['Total Users:', total_users])
        ws.append(['Cost per User:', f"{billing_config['currency']} {cost_per_user:.2f}"])
        ws.append(['Total Monthly Cost:', f"{billing_config['currency']} {total_cost:.2f}"])
        ws['A7'].font = Font(bold=True)
        
        # Empty row
        ws.append([])
        
        # User details header
        headers = ['Email', 'Created Date', 'Last Login', 'IMEI', 'IMEI Added Date', 'Total IMEIs']
        ws.append(headers)
        
        # Style header row
        for cell in ws[9]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # User data
        for email, user_data in sorted(users_db.items()):
            imeis = user_imeis.get(email, [])
            
            if imeis:
                # First row with first IMEI
                created_date = datetime.fromisoformat(user_data['created_at']).strftime('%Y-%m-%d %H:%M')
                last_login = datetime.fromisoformat(user_data['last_login']).strftime('%Y-%m-%d %H:%M')
                
                for idx, imei_data in enumerate(imeis):
                    imei_added = datetime.fromisoformat(imei_data['added_date']).strftime('%Y-%m-%d %H:%M')
                    
                    if idx == 0:
                        # First IMEI row includes user details
                        ws.append([
                            email,
                            created_date,
                            last_login,
                            imei_data['imei'],
                            imei_added,
                            len(imeis)
                        ])
                    else:
                        # Subsequent IMEIs
                        ws.append([
                            '',
                            '',
                            '',
                            imei_data['imei'],
                            imei_added,
                            ''
                        ])
            else:
                # User with no IMEIs
                created_date = datetime.fromisoformat(user_data['created_at']).strftime('%Y-%m-%d %H:%M')
                last_login = datetime.fromisoformat(user_data['last_login']).strftime('%Y-%m-%d %H:%M')
                ws.append([
                    email,
                    created_date,
                    last_login,
                    'No IMEIs',
                    '',
                    0
                ])
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=billing_statement_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except ImportError:
        return jsonify({
            'error': 'openpyxl library not installed. Run: pip install openpyxl'
        }), 500
    except Exception as e:
        if DEBUG_MODE:
            print(f"❌ Error in admin_export_billing: {str(e)}")
        return jsonify({'error': f'Internal server error: {str(e)}'}), 500


# Serve admin pages
@app.route('/admin')
def serve_admin_login():
    """Serve admin login page"""
    return send_from_directory('../', 'admin-login.html')


@app.route('/admin/dashboard')
def serve_admin_dashboard():
    """Serve admin dashboard page"""
    return send_from_directory('../', 'admin-dashboard.html')


if __name__ == '__main__':
    if DEBUG_MODE:
        print("=" * 60)
        print("🚀 BLE Tag Tracker Backend Server")
        print("=" * 60)
        print(f"📱 Frontend: http://{HOST if HOST != '0.0.0.0' else 'localhost'}:{PORT}")
        print(f"🔑 Token API: http://{HOST if HOST != '0.0.0.0' else 'localhost'}:{PORT}/api/token")
        print(f"❤️  Health Check: http://{HOST if HOST != '0.0.0.0' else 'localhost'}:{PORT}/api/health")
        print(f"🔧 Mode: {'DEBUG' if DEBUG_MODE else 'PRODUCTION'}")
        print("=" * 60)
        print("✅ Server starting...")
        print()
    
    app.run(
        host=HOST,
        port=PORT,
        debug=DEBUG_MODE
    )
